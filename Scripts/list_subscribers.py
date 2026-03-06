#!/usr/bin/env python3
"""
list_subscribers.py
-------------------
Fetches all subscribers from Supabase and saves to an Excel file.
Run from anywhere on your Mac:

    python3 /Users/ncm/Pictures/Mohangraphy/Scripts/list_subscribers.py
"""

import json
import urllib.request
import urllib.error
import os
import subprocess
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Installing openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load credentials
CONTENT_FILE = os.path.join(os.path.dirname(__file__), 'content.json')
try:
    with open(CONTENT_FILE, 'r', encoding='utf-8') as f:
        config = json.load(f)
except FileNotFoundError:
    print(f"content.json not found at: {CONTENT_FILE}")
    raise SystemExit(1)

site     = config.get('site', {})
SUPA_URL = site.get('supabase_url', 'https://xjcpryfgodgqqtbblklg.supabase.co')
SUPA_KEY = site.get('supabase_service_role_key',
           site.get('supabase_anon_key',
           site.get('supabase_key',
           'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InhqY3ByeWZnb2RncXF0YmJsa2xnIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzIxODEzMjcsImV4cCI6MjA4Nzc1NzMyN30.M9KoprG4uaH3wcZ7nI0Hip4IAdqiy8m5UoiB9DzjreI')))

# Fetch subscribers
url = f"{SUPA_URL}/rest/v1/subscribers?select=name,email,subscribed_at&order=subscribed_at.asc"
req = urllib.request.Request(url, headers={
    'apikey': SUPA_KEY,
    'Authorization': f'Bearer {SUPA_KEY}',
    'Content-Type': 'application/json',
})
try:
    with urllib.request.urlopen(req) as resp:
        subscribers = json.loads(resp.read().decode('utf-8'))
except urllib.error.HTTPError as e:
    print(f"Supabase error {e.code}: {e.read().decode()}")
    raise SystemExit(1)
except urllib.error.URLError as e:
    print(f"Network error: {e.reason}")
    raise SystemExit(1)

# Build Excel
wb = Workbook()
ws = wb.active
ws.title = "Subscribers"

GOLD      = 'C9A96E'
DARK      = '1A1A1A'
LT_GREY   = 'F5F5F5'
WHITE     = 'FFFFFF'
MED_GREY  = 'CCCCCC'

def border():
    s = Side(style='thin', color=MED_GREY)
    return Border(left=s, right=s, top=s, bottom=s)

# Title
ws.merge_cells('A1:D1')
ws['A1'] = 'MOHANGRAPHY — SUBSCRIBERS'
ws['A1'].font      = Font(name='Arial', bold=True, size=14, color=WHITE)
ws['A1'].fill      = PatternFill('solid', fgColor=DARK)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

# Subtitle
ws.merge_cells('A2:D2')
ws['A2'] = f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}   |   Total: {len(subscribers)} subscriber(s)"
ws['A2'].font      = Font(name='Arial', size=10, color='888888')
ws['A2'].fill      = PatternFill('solid', fgColor='2A2A2A')
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 20

# Headers
for col, h in enumerate(['#', 'Name', 'Email', 'Subscribed On'], 1):
    c = ws.cell(row=3, column=col, value=h)
    c.font      = Font(name='Arial', bold=True, size=11, color=WHITE)
    c.fill      = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border    = border()
ws.row_dimensions[3].height = 22

# Data
for i, s in enumerate(subscribers, 1):
    name   = (s.get('name')  or '—').strip()
    email  = (s.get('email') or '—').strip()
    raw_dt = s.get('subscribed_at', '')
    try:
        date_str = datetime.fromisoformat(raw_dt.replace('Z', '+00:00')).strftime('%d %b %Y')
    except Exception:
        date_str = raw_dt[:10] if raw_dt else '—'

    fill = PatternFill('solid', fgColor=LT_GREY if i % 2 == 0 else WHITE)
    for col, (val, align) in enumerate(zip([i, name, email, date_str],
                                           ['center','left','left','center']), 1):
        c = ws.cell(row=i+3, column=col, value=val)
        c.font      = Font(name='Arial', size=11, color=DARK)
        c.fill      = fill
        c.alignment = Alignment(horizontal=align, vertical='center')
        c.border    = border()
    ws.row_dimensions[i+3].height = 20

# Column widths & freeze
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 36
ws.column_dimensions['D'].width = 18
ws.freeze_panes = 'A4'

# Save
out = os.path.join(os.path.dirname(__file__),
      f"subscribers_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
wb.save(out)
print(f"\n  Saved {len(subscribers)} subscriber(s) to:\n   {out}\n")
